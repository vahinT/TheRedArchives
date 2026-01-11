import os
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinterdnd2 import DND_FILES, TkinterDnD

def find_last_row(sheet):
    for row in range(sheet.max_row, 0, -1):
        if any(cell.value is not None for cell in sheet[row]):
            return row
    return 0

def copy_sheet_data(source_sheet, target_sheet, start_row):
    for i, row in enumerate(source_sheet.iter_rows(values_only=False), start=start_row):
        for j, cell in enumerate(row, start=1):
            target_sheet.cell(row=i, column=j, value=cell.value)

def get_unique_output_name(base="output.xlsx"):
    if not os.path.exists(base):
        return base
    name, ext = os.path.splitext(base)
    i = 1
    while os.path.exists(f"{name} ({i}){ext}"):
        i += 1
    return f"{name} ({i}){ext}"

def merge_excel_files(file_paths, rows_gap):
    if len(file_paths) < 2:
        messagebox.showerror("Error", "Please add at least 2 Excel files.")
        return

    workbooks = [openpyxl.load_workbook(path) for path in file_paths]
    main_wb = workbooks[0]

    for wb in workbooks[1:]:
        for sheet_name in wb.sheetnames:
            if sheet_name in main_wb.sheetnames:
                target_sheet = main_wb[sheet_name]
                source_sheet = wb[sheet_name]
                last_row = find_last_row(target_sheet)
                paste_row = last_row + rows_gap + 1
                copy_sheet_data(source_sheet, target_sheet, paste_row)

    output_name = get_unique_output_name("output.xlsx")
    main_wb.save(output_name)
    messagebox.showinfo("Success", f"✅ Merged workbook saved as:\n{output_name}")

def on_drop(event):
    files = root.tk.splitlist(event.data)
    for f in files:
        if f.endswith(".xlsx") and f not in file_list:
            file_list.append(f)
            filebox.insert(tk.END, os.path.basename(f))

def add_files():
    files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    for f in files:
        if f.endswith(".xlsx") and f not in file_list:
            file_list.append(f)
            filebox.insert(tk.END, os.path.basename(f))

def merge_action():
    try:
        gap = int(gap_entry.get())
        if gap < 0:
            raise ValueError
    except ValueError:
        gap = 3  # fallback default
        messagebox.showwarning("Warning", "Invalid number of rows. Defaulting to 3.")
    merge_excel_files(file_list, gap)

def clear_files():
    file_list.clear()
    filebox.delete(0, tk.END)

# GUI Setup
root = TkinterDnD.Tk()
root.title("Vahin's Workbook Merger!")
root.geometry("520x450")
root.config(bg="#f0f0f0")

file_list = []

tk.Label(root, text="Drag and drop Excel files here or click 'Add Files'", bg="#f0f0f0", font=("Arial", 12)).pack(pady=10)

filebox = tk.Listbox(root, width=60, height=10)
filebox.pack(pady=10)
filebox.drop_target_register(DND_FILES)
filebox.dnd_bind('<<Drop>>', on_drop)

gap_frame = tk.Frame(root, bg="#f0f0f0")
gap_frame.pack(pady=5)
tk.Label(gap_frame, text="Blank Rows Between Merges:", bg="#f0f0f0").grid(row=0, column=0, padx=5)
gap_entry = tk.Entry(gap_frame, width=5)
gap_entry.insert(0, "3")
gap_entry.grid(row=0, column=1)

btn_frame = tk.Frame(root, bg="#f0f0f0")
btn_frame.pack(pady=10)

tk.Button(btn_frame, text="Add Files", command=add_files, width=12).grid(row=0, column=0, padx=5)
tk.Button(btn_frame, text="Clear Files", command=clear_files, width=12).grid(row=0, column=1, padx=5)
tk.Button(btn_frame, text="Merge Files", command=merge_action, width=12, bg="#4CAF50", fg="white").grid(row=0, column=2, padx=5)

root.mainloop()
